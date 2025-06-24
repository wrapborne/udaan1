# admin_app.py
import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime, date
from extract_premium_summary import extract_from_pdf, extract_from_txt
from extractor import extract_all_lic_data
from utils import get_mysql_connection, get_policy_count_by_plan, load_lic_data_from_db, filter_df_by_selected_year, filter_df_by_financial_year
from db_utils import get_pending_users, add_user, delete_pending_user, get_all_users, update_user_role_and_start, delete_user
from openpyxl import load_workbook
from openpyxl.styles import numbers
from data_display_column import ADMIN_DISPLAY_COLUMNS as DISPLAY_COLUMNS


# --- File Upload Handlers ---
def upload_lic_data(uploaded_file):
    if "db_name" not in st.session_state:
        st.error("🔒 Please log in first.")
        return

    # 1. Save uploaded file temporarily
    with open("latest_uploaded.txt", "wb") as f:
        f.write(uploaded_file.getbuffer())

    # 2. Extract data using your custom logic
    new_data = extract_all_lic_data("latest_uploaded.txt")
    if new_data.empty:
        st.warning("⚠️ No valid data found in uploaded file.")
        return

    db_name = st.session_state["db_name"]
    engine = get_mysql_connection(db_name)

    try:
        # 3. Read existing data (if table exists)
        try:
            old_data = pd.read_sql("SELECT * FROM lic_data", con=engine)
        except Exception:
            old_data = pd.DataFrame()

        # 4. Combine and deduplicate
        combined = pd.concat([old_data, new_data]).drop_duplicates(subset=["Policy No"], keep="last")

        # Optional: Add uploader info
        if "uploaded_by" not in combined.columns:
            combined["uploaded_by"] = st.session_state["username"]

        # 5. Save to MySQL (full overwrite — OK if only 1 agent per DB)
        combined.to_sql("lic_data", con=engine, if_exists="replace", index=False)

        st.success(f"✅ {len(combined)} proposals saved successfully.")

    except Exception as e:
        st.error(f"❌ Failed to save data: {e}")

    finally:
        # 6. Cleanup file
        if os.path.exists("latest_uploaded.txt"):
            os.remove("latest_uploaded.txt")

def upload_premium_summary(premium_file):
    if "db_name" not in st.session_state:
        st.error("🔒 Login required. Please log in again.")
        st.stop()

    try:
        # 1. Save uploaded file temporarily
        with open("temp_premium_upload", "wb") as f:
            f.write(premium_file.getbuffer())

        # 2. Extract data based on file type
        if premium_file.name.endswith(".pdf"):
            premium_df = extract_from_pdf("temp_premium_upload")
        else:
            premium_df = extract_from_txt("temp_premium_upload")

        # 3. Clean and prepare data
        premium_df.rename(columns={"Report Month": "report_month"}, inplace=True)
        premium_df["agency_code"] = premium_df["Agency Code"].str.strip().str.upper()
        premium_df["uploaded_by"] = st.session_state.username

        # 4. Connect to correct DO database
        engine = get_mysql_connection(st.session_state["db_name"])
        report_month = premium_df["report_month"].iloc[0]

        # 5. Check if data already exists for this month + uploader
        existing = pd.read_sql(
            "SELECT * FROM premium_summary WHERE report_month = %s AND uploaded_by = %s",
            engine,
            params=(report_month, st.session_state.username)
        )

        check_cols = ["agency_code", "report_month", "total_premium", "fp_sch_prem", "fy_sch_prem", "uploaded_by"]

        if not existing.empty:
            # Compare old vs new
            existing_sorted = existing[check_cols].sort_values(by=check_cols).reset_index(drop=True)
            new_sorted = premium_df[check_cols].sort_values(by=check_cols).reset_index(drop=True)

            if existing_sorted.equals(new_sorted):
                # Exact match — overwrite
                with engine.begin() as conn:
                    conn.execute(
                        "DELETE FROM premium_summary WHERE report_month = %s AND uploaded_by = %s",
                        (report_month, st.session_state.username)
                    )
                premium_df[check_cols].to_sql("premium_summary", con=engine, if_exists="append", index=False)
                st.success(f"♻️ Existing data for {report_month} matched — overwritten.")
            else:
                # Conflict — show warning and checkbox
                st.warning(f"⚠️ Data for {report_month} already exists and does not match. Upload skipped.")

                if st.checkbox(f"☑️ Force overwrite data for {report_month}", key=f"override_{report_month}"):
                    from sqlalchemy import text
                    with engine.connect() as conn:
                        delete_stmt = text("DELETE FROM premium_summary WHERE report_month = :rm AND uploaded_by = :up")
                        conn.execute(delete_stmt, {"rm": report_month, "up": st.session_state.username})

                    premium_df[check_cols].to_sql("premium_summary", con=engine, if_exists="append", index=False)
                    st.success(f"✅ Data for {report_month} forcibly overwritten.")
        else:
            # New data — insert directly
            premium_df[check_cols].to_sql("premium_summary", con=engine, if_exists="append", index=False)
            st.success("✅ Premium summary uploaded and saved.")
            
    except Exception as e:
        st.error(f"❌ Failed to process premium file: {e}")


# --- Premium Summary Dropdown ---
def show_premium_summary_dropdown():
    with st.expander("📈 View Premium Summary by Month"):
        try:
            engine = get_mysql_connection(st.session_state.db_name)
            username = st.session_state.username
            query = f"SELECT * FROM premium_summary WHERE uploaded_by = '{username}'"
            df = pd.read_sql(query, engine)
            
            if df.empty:
                st.info("No premium summary uploaded yet.")
                return

            unique_months = df["report_month"].dropna().unique()
            selected_month = st.selectbox("🗓️ Select Report Month", sorted(unique_months, reverse=True))
            month_df = df[df["report_month"] == selected_month].copy()
            eligible_premium_sum = month_df["fp_sch_prem"].sum() + month_df["fy_sch_prem"].sum()
            st.dataframe(month_df, use_container_width=True)
            st.success(f"🧮 Total Eligible Premium for {selected_month}: ₹{eligible_premium_sum:,.2f}")

        except Exception as e:
            st.warning(f"⚠️ Could not load premium summary: {e}")


# --- Pending and Approved Users ---
def show_pending_approvals():
    st.markdown("---")
    if st.button("📋 View Pending Registrations", use_container_width=True):
        st.session_state.show_pending = not st.session_state.get("show_pending", False)

    if st.session_state.get("show_pending"):
        st.subheader("👥 Pending Agent Approvals")
        pending = get_pending_users()
        if not pending:
            st.info("✅ No pending agent registrations.")
        else:
            for row in pending:
                # ✅ Backward-compatible unpacking
                rowid = row[0]
                username = row[1]
                password = row[2]
                role = row[3]
                admin_username = row[4]
                db_name = row[5]
                do_code = row[6] if len(row) > 6 else None
                agency_code = row[7] if len(row) > 7 else None

                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(f"🔸 **{username}** | Role: `{role}`")
                with col2:
                    if st.button("✅ Approve", key=f"approve_{rowid}"):
                        add_user(
                            username=username,
                            password=password,
                            role=role,
                            start_date=datetime.today(),
                            admin_username=admin_username,
                            db_name=db_name,
                            do_code=do_code,
                            agency_code=agency_code
                        )
                        delete_pending_user(rowid)
                        st.success(f"User `{username}` approved.")
                        st.rerun()


def show_user_management():
    st.markdown("## 👥 Manage Registered Users")
    users = get_all_users()
    if not users:
        st.info("No registered users found.")
        return
  #  full_name = st.session_state.get("name", st.session_state.get("username", "User")).title()
    for username, role, start_date, do_code in users:
        if username == st.session_state.username:
            continue  # 🔒 Skip logged-in user
        if st.session_state.role != "superadmin" and role == "superadmin":
            continue  # 🔒 Hide superadmin from admins
        with st.expander(f"🔸 {username}"):
            st.text(f"DO Code: {do_code if do_code else 'N/A'}")
            if role == "superadmin":
                st.markdown("🔒 Superadmin role (cannot modify)")
                continue

            st.text("Role: " + role)  # 🔒 Superadmin can't change roles

            if isinstance(start_date, str):
                parsed_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            elif isinstance(start_date, datetime):
                parsed_date = start_date.date()
            else:
                parsed_date = start_date if isinstance(start_date, date) else datetime.today().date()

            if isinstance(start_date, str):
                start_date_val = datetime.strptime(start_date, "%Y-%m-%d")
            else:
                start_date_val = start_date or datetime.today()

            new_start_date = st.date_input("Start Date", value=start_date_val, key=f"date_{username}")

            cols = st.columns(2)
            with cols[0]:
                if st.button("💾 Save Changes", key=f"save_{username}"):
                    update_user_role_and_start(username, new_role, new_start_date.strftime("%Y-%m-%d"))
                    st.success("✅ User updated.")
                    st.rerun()
            with cols[1]:
                if role.lower() != "admin":
                    if st.checkbox(f"⚠️ Confirm delete {username}", key=f"confirm_delete_{username}"):
                        if st.button("🗑️ Delete User", key=f"delete_{username}"):
                            delete_user(username)
                            st.warning(f"❌ User `{username}` deleted.")
                            st.rerun()
                else:
                    st.caption("🔒 Admins can't be deleted.")


# --- Data Display and Filtering ---
def show_agent_data(df):
    if df.empty:
        st.warning("No data to display.")
        return

    df = df.copy().reset_index(drop=True)
    df.insert(0, "S.No.", range(1, len(df) + 1))
    df["DOC"] = pd.to_datetime(df["DOC"], errors="coerce")

    # Ensure DOC is parsed safely
    df["DOC"] = pd.to_datetime(df["DOC"], errors="coerce")
    
    min_doc, max_doc = df["DOC"].min(), df["DOC"].max()

    # Fallback to today's date if invalid
    if pd.isnull(min_doc) or pd.isnull(max_doc):
        min_doc = max_doc = datetime.today()
        
    date_range = st.date_input("🗓️ Filter by DOC", value=(min_doc, max_doc), key="date_filter")
    if isinstance(date_range, tuple) and len(date_range) == 2:
        df = df[(df["DOC"] >= pd.to_datetime(date_range[0])) & (df["DOC"] <= pd.to_datetime(date_range[1]))]

    plan_options = ["All Plans"] + sorted(df["Plan"].dropna().astype(str).unique().tolist())
    selected_plan = st.selectbox("📋 Filter by Plan", plan_options, key="plan_filter")
    if selected_plan != "All Plans":
        df = df[df["Plan"].astype(str) == selected_plan]

    mode_options = ["All Modes"] + sorted(df["Mode"].dropna().astype(str).unique().tolist())
    selected_mode = st.selectbox("💼 Filter by Mode", mode_options, key="mode_filter")
    if selected_mode != "All Modes":
        df = df[df["Mode"].astype(str) == selected_mode]

    search = st.text_input("🔍 Search", key="search_filter")
    if search:
        df = df[df.apply(lambda row: search.lower() in str(row).lower(), axis=1)]

    st.info(f"🔢 Total Proposals: {len(df)} | 🟢 ANANDA: {(df['ANANDA'].str.strip().str.upper() == 'YES').sum()}")

    df["DOC"] = df["DOC"].dt.strftime("%d/%m/%Y")
    
    display_cols = [col for col in DISPLAY_COLUMNS if col in df.columns]
    st.dataframe(df[display_cols], use_container_width=True)

    with io.BytesIO() as buffer:
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "Policy No":
                for cell in col[1:]:
                    cell.number_format = numbers.FORMAT_TEXT
        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)
        st.download_button("📅 Download Data", data=final_buffer.getvalue(), file_name="LIC_Data.xlsx")

    st.markdown("### 📊 Policy Count by Plan")
    plan_count_df = get_policy_count_by_plan(df)
    if not plan_count_df.empty:
        st.dataframe(plan_count_df.style.highlight_max(axis=1), use_container_width=True)

    show_pending_approvals()
    show_user_management()


# --- Main Dashboard ---
def admin_dashboard():
    st.title("🧑‍💼 Admin Panel")
    show_premium_summary_dropdown()

    df = load_lic_data_from_db()
    df = filter_df_by_selected_year(df, st.session_state.get("selected_year", "All Years"))
    df = filter_df_by_financial_year(df, st.session_state.get("fin_year", "All Financial Years"))
    show_agent_data(df)

    st.markdown("---")
    st.markdown("### 📤 Upload Files")
    col1, col2 = st.columns(2)
    with col1:
        uploaded = st.file_uploader("📄 Upload LIC Proposal Register (.txt)", type="txt")
        if uploaded and not st.session_state.get("data_uploaded", False):
            upload_lic_data(uploaded)
            st.session_state["data_uploaded"] = True
            st.rerun()
        if st.session_state.get("data_uploaded", False):
            st.success("✅ Data uploaded and updated.")
            st.session_state["data_uploaded"] = False  # Reset the flag
    


    with col2:
        premium_file = st.file_uploader("💰 Upload Premium Summary (PDF or TXT)", type=["pdf", "txt"], key="premium_uploader")

        if premium_file and not st.session_state.get("premium_uploaded", False):
            upload_premium_summary(premium_file)
            st.session_state["premium_uploaded"] = True
            st.rerun()

        # ✅ Show success message once after rerun
        if st.session_state.get("premium_uploaded", False):
            st.success("✅ Premium summary uploaded and saved.")
            st.session_state["premium_uploaded"] = False
