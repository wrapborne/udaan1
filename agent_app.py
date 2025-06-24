# agent_app.py
import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
from utils import load_lic_data_from_db, filter_df_by_selected_year, filter_df_by_financial_year, get_policy_count_by_plan
from data_display_column import AGENT_DISPLAY_COLUMNS as DISPLAY_COLUMNS


def show_agent_data(df):
    if df.empty:
        st.warning("No data to display.")
        return

    df = df.copy().reset_index(drop=True)
    df.insert(0, "S.No.", range(1, len(df) + 1))
    df["DOC"] = pd.to_datetime(df["DOC"], errors="coerce")

    min_doc, max_doc = df["DOC"].min(), df["DOC"].max()
    date_range = st.date_input("🗓️ Filter by DOC", value=(min_doc, max_doc))
    if isinstance(date_range, tuple) and len(date_range) == 2:
        df = df[(df["DOC"] >= pd.to_datetime(date_range[0])) & (df["DOC"] <= pd.to_datetime(date_range[1]))]

    plan_options = ["All Plans"] + sorted(df["Plan"].dropna().astype(str).unique().tolist())
    selected_plan = st.selectbox("📋 Filter by Plan", plan_options)
    if selected_plan != "All Plans":
        df = df[df["Plan"].astype(str) == selected_plan]

    mode_options = ["All Modes"] + sorted(df["Mode"].dropna().astype(str).unique().tolist())
    selected_mode = st.selectbox("💼 Filter by Mode", mode_options)
    if selected_mode != "All Modes":
        df = df[df["Mode"].astype(str) == selected_mode]

    search = st.text_input("🔍 Search")
    if search:
        df = df[df.apply(lambda row: search.lower() in str(row).lower(), axis=1)]

    st.info(f"🔢 Total Proposals: {len(df)} | 🟢 ANANDA: {(df['ANANDA'].str.strip().str.upper() == 'YES').sum()}")

    df["DOC"] = df["DOC"].dt.strftime("%d/%m/%Y")

    display_cols = [col for col in DISPLAY_COLUMNS if col in df.columns]
    st.dataframe(df[display_cols], use_container_width=True)

    # Excel export
    with io.BytesIO() as buffer:
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "Policy No":
                for cell in col[1:]:
                    cell.number_format = numbers.FORMAT_TEXT
        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)
        st.download_button("📅 Download Data", data=final_buffer.getvalue(), file_name="Agent_Data.xlsx")

    st.markdown("### 📊 Policy Count by Plan")
    plan_count_df = get_policy_count_by_plan(df)
    if not plan_count_df.empty:
        st.dataframe(plan_count_df.style.highlight_max(axis=1), use_container_width=True)


def agent_dashboard():
    st.title("📋 My Agency")

    df = load_lic_data_from_db()

    # ✅ Get agency_code from session
    agency_code = st.session_state.get("agency_code")
    if not agency_code:
            st.error("❌ No agency code found for this user.")
            return
    # ✅ Filter using agency_code
    df["Agency Code"] = df["Agency Code"].astype(str)
    df = df[df["Agency Code"].str.strip().str.upper() == agency_code.strip().upper()]

    df = filter_df_by_selected_year(df, st.session_state.get("selected_year", "All Years"))
    df = filter_df_by_financial_year(df, st.session_state.get("fin_year", "All Financial Years"))

    show_agent_data(df)

